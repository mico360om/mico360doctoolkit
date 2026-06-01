"""v5.4 feature tests: new dedicated PDF tools, page numbers, sign, metadata,
searchable OCR, Office conversions (engine-gated), and image Resize/Convert/Watermark.

Run:  python tests/v54_features_test.py
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

failures: list[str] = []
rep = lambda m: None  # noqa: E731


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"[{'PASS' if ok else 'FAIL'}] {name}" + (f": {detail}" if detail else ""))
    if not ok:
        failures.append(name)


def _make_pdf(path: Path, pages: int) -> Path:
    import fitz
    doc = fitz.open()
    for i in range(pages):
        pg = doc.new_page(width=400, height=300)
        pg.insert_text((50, 80), f"PAGE {i + 1}", fontsize=24)
    doc.save(str(path)); doc.close()
    return path


def _scanned_pdf(path: Path) -> Path:
    import fitz
    src = fitz.open()
    pg = src.new_page(width=500, height=300)
    pg.insert_text((50, 120), "Searchable Test Document", fontsize=24)
    pix = pg.get_pixmap(dpi=150); src.close()
    img = path.with_suffix(".png"); pix.save(str(img))
    flat = fitz.open(); fp = flat.new_page(width=500, height=300)
    fp.insert_image(fp.rect, filename=str(img)); flat.save(str(path)); flat.close()
    return path


def main() -> int:
    from PySide6.QtWidgets import QApplication
    if QApplication.instance() is None:
        QApplication([])
    from mico360.core import processors as P
    import fitz

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td); out = tmp / "out"; out.mkdir()
        src = _make_pdf(tmp / "doc.pdf", 6)

        # --- Rotate / Delete / Extract ----------------------------------
        r = P.pdf_rotate(src, out, {"angle": 90, "pages": "2", "overwrite": True}, rep)
        d = fitz.open(str(r[0])); check("rotate: page 2 rotated",
                                        d[1].rotation == 90 and d[0].rotation == 0); d.close()
        r = P.pdf_delete(src, out, {"pages": "2,4", "overwrite": True}, rep)
        d = fitz.open(str(r[0])); check("delete: 6-2=4 pages", d.page_count == 4); d.close()
        r = P.pdf_extract(src, out, {"pages": "5,1", "overwrite": True}, rep)
        d = fitz.open(str(r[0]))
        check("extract: order 5 then 1",
              d[0].get_text().strip() == "PAGE 5" and d[1].get_text().strip() == "PAGE 1")
        d.close()

        # --- Page numbers ------------------------------------------------
        r = P.pdf_page_numbers(src, out, {"position": "bottom-center",
                                          "format": "n_of_total", "start": 1,
                                          "overwrite": True}, rep)
        d = fitz.open(str(r[0])); txt = d[0].get_text(); d.close()
        check("page numbers: '1 / 6' stamped", "1 / 6" in txt, repr(txt[-20:]))

        # --- Metadata ----------------------------------------------------
        r = P.pdf_metadata(src, out, {"title": "My Title", "author": "Me",
                                      "overwrite": True}, rep)
        from pypdf import PdfReader
        meta = PdfReader(str(r[0])).metadata
        check("metadata: title set", (meta or {}).get("/Title") == "My Title")
        check("metadata: author set", (meta or {}).get("/Author") == "Me")

        # --- Sign (signature image) -------------------------------------
        from PIL import Image
        sig = tmp / "sig.png"; Image.new("RGBA", (160, 60), (10, 10, 200, 255)).save(sig)
        before = len(fitz.open(str(src))[-1].get_images())
        r = P.pdf_sign(src, out, {"image_path": str(sig), "page": "last",
                                  "position": "bottom-right", "width": 30,
                                  "overwrite": True}, rep)
        d = fitz.open(str(r[0])); after = len(d[-1].get_images()); npg = d.page_count; d.close()
        check("sign: image added to last page", after > before, f"{before}->{after}")
        check("sign: page count preserved", npg == 6)

        # --- Searchable OCR ---------------------------------------------
        try:
            scanned = _scanned_pdf(tmp / "scan.pdf")
            had_text = P._page_has_text(fitz.open(str(scanned))[0])
            r = P.pdf_ocr(scanned, out, {"overwrite": True}, rep)
            d = fitz.open(str(r[0])); now_text = d[0].get_text().strip(); d.close()
            check("searchable OCR: scanned page had no text before", not had_text)
            check("searchable OCR: text layer added", "Searchable" in now_text or
                  len(now_text) > 0, repr(now_text[:40]))
        except P.ProcessError as exc:
            print(f"[INFO] OCR unavailable, skipping: {exc}")

        # --- PDF -> Excel (text fallback) -------------------------------
        r = P.pdf_to_excel(src, out, {"overwrite": True}, rep)
        from openpyxl import load_workbook
        wb = load_workbook(str(r[0]))
        check("pdf->excel: produced an .xlsx with sheets", len(wb.sheetnames) >= 1,
              str(wb.sheetnames))

        # --- Excel -> PDF / PPT -> PDF (engine-gated) -------------------
        from mico360.core.deps import find_libreoffice
        lo = find_libreoffice()
        from openpyxl import Workbook
        xlsx = tmp / "book.xlsx"; w = Workbook(); ws = w.active
        ws.append(["A", "B"]); ws.append([1, 2]); w.save(str(xlsx))
        try:
            r = P.excel_to_pdf(xlsx, out, {"overwrite": True}, rep)
            check("excel->pdf produced a PDF", r and r[0].exists())
        except P.ProcessError as exc:
            check("excel->pdf gives a clear message without an engine",
                  "LibreOffice" in str(exc), str(exc)[:60])
            print(f"[INFO] no Office engine (LibreOffice={bool(lo)}); excel->pdf skipped")

        # --- Image Resize / Convert / Watermark -------------------------
        photo = tmp / "photo.png"
        Image.new("RGB", (800, 600), (90, 140, 200)).save(photo)
        r = P.image_resize(photo, out, {"mode": "dimensions", "width": 400,
                                        "keep_aspect": True, "overwrite": True}, rep)
        im = Image.open(r[0]); check("resize: 800x600 -> 400x300",
                                     im.size == (400, 300), str(im.size)); im.close()
        r = P.image_resize(photo, out, {"mode": "percent", "percent": 25,
                                        "overwrite": True}, rep)
        im = Image.open(r[0]); check("resize percent: -> 200x150", im.size == (200, 150),
                                     str(im.size)); im.close()
        r = P.image_convert(photo, out, {"format": "jpg", "quality": 85,
                                         "overwrite": True}, rep)
        check("convert: png -> jpg", r[0].suffix.lower() == ".jpg")
        r = P.image_watermark(photo, out, {"wm_type": "text", "text": "DRAFT",
                                           "opacity": 35, "font_size": 60,
                                           "color": "red", "rotation": 30,
                                           "overwrite": True}, rep)
        check("image watermark (text): output exists", r[0].exists())
        logo = tmp / "logo.png"; Image.new("RGBA", (200, 100), (0, 150, 0, 255)).save(logo)
        r = P.image_watermark(photo, out, {"wm_type": "image", "image_path": str(logo),
                                           "scale": 30, "opacity": 50,
                                           "overwrite": True}, rep)
        check("image watermark (logo): output exists", r[0].exists())

    print()
    if failures:
        print(f"{len(failures)} check(s) failed: {', '.join(failures)}")
        return 1
    print("All v5.4 feature checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
