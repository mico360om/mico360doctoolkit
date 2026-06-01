"""Edge-case / error-handling tests for v5.4 tools, plus real PDF->Excel tables.

Run:  python tests/v54_edge_test.py
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

failures: list[str] = []
rep = lambda m: None  # noqa: E731


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"[{'PASS' if ok else 'FAIL'}] {name}" + (f": {detail}" if detail else ""))
    if not ok:
        failures.append(name)


def raises(fn) -> bool:
    from mico360.core.processors import ProcessError
    try:
        fn()
        return False
    except ProcessError:
        return True


def main() -> int:
    from PySide6.QtWidgets import QApplication
    if QApplication.instance() is None:
        QApplication([])
    from mico360.core import processors as P
    import fitz

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td); out = tmp / "out"; out.mkdir()

        def pdf(n=3):
            p = tmp / f"d{n}.pdf"
            doc = fitz.open()
            for i in range(n):
                doc.new_page(width=300, height=200)
            doc.save(str(p)); doc.close()
            return p

        src = pdf(3)
        o = {"overwrite": True}

        # --- error handling -------------------------------------------------
        check("delete empty -> error",
              raises(lambda: P.pdf_delete(src, out, {**o, "pages": ""}, rep)))
        check("delete all -> error",
              raises(lambda: P.pdf_delete(src, out, {**o, "pages": "1-3"}, rep)))
        check("extract invalid -> error",
              raises(lambda: P.pdf_extract(src, out, {**o, "pages": "99,abc"}, rep)))
        check("rotate non-90 -> error",
              raises(lambda: P.pdf_rotate(src, out, {**o, "angle": 45, "pages": "all"}, rep)))
        check("sign missing image -> error",
              raises(lambda: P.pdf_sign(src, out, {**o, "image_path": ""}, rep)))
        check("sign bad path -> error",
              raises(lambda: P.pdf_sign(src, out, {**o, "image_path": str(tmp / "no.png")}, rep)))
        check("resize no dimensions -> error",
              raises(lambda: P.image_resize(_png(tmp), out,
                     {**o, "mode": "dimensions", "width": 0, "height": 0}, rep)))
        check("image watermark missing logo -> error",
              raises(lambda: P.image_watermark(_png(tmp), out,
                     {**o, "wm_type": "image", "image_path": ""}, rep)))
        check("metadata all-empty -> keeps file (no error)",
              not raises(lambda: P.pdf_metadata(src, out, {**o}, rep)))

        # --- corrupt input is reported, not crashed -------------------------
        bad = tmp / "bad.pdf"; bad.write_bytes(b"%PDF-1.4 not really")
        check("rotate corrupt pdf -> friendly error",
              raises(lambda: P.pdf_rotate(bad, out, {**o, "pages": "all"}, rep)))

        # --- real table extraction (PDF -> Excel) ---------------------------
        tbl = tmp / "table.pdf"
        doc = fitz.open(); pg = doc.new_page(width=400, height=300)
        shape = pg.new_shape()
        for x in (50, 160, 270, 360):
            shape.draw_line((x, 60), (x, 150))
        for y in (60, 105, 150):
            shape.draw_line((50, y), (360, y))
        shape.finish(width=1, color=(0, 0, 0)); shape.commit()
        cells = [("Name", 60, 90), ("Age", 175, 90), ("City", 285, 90),
                 ("Ann", 60, 135), ("30", 175, 135), ("NYC", 285, 135)]
        for t, x, y in cells:
            pg.insert_text((x, y), t, fontsize=11)
        doc.save(str(tbl)); doc.close()

        r = P.pdf_to_excel(tbl, out, o, rep)
        from openpyxl import load_workbook
        wb = load_workbook(str(r[0]))
        all_text = " ".join(str(c.value) for ws in wb.worksheets
                            for row in ws.iter_rows() for c in row if c.value)
        check("pdf->excel extracts the table cells",
              "Name" in all_text and "Ann" in all_text and "NYC" in all_text,
              all_text[:60])
        check("pdf->excel made a table sheet (P*_T*)",
              any(s.startswith("P1_T") for s in wb.sheetnames), str(wb.sheetnames))

    print()
    if failures:
        print(f"{len(failures)} check(s) failed: {', '.join(failures)}")
        return 1
    print("All v5.4 edge-case checks passed.")
    return 0


def _png(tmp: Path) -> Path:
    from PIL import Image
    p = tmp / "img.png"
    Image.new("RGB", (300, 200), (10, 20, 30)).save(p)
    return p


if __name__ == "__main__":
    raise SystemExit(main())
